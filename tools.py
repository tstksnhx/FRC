import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
import torch
from sklearn import metrics
from sklearn.preprocessing import StandardScaler
from torch.utils.data import Dataset
import mathtool as mt
import data_load_tools as dlt
import random
from sklearn.calibration import CalibratedClassifierCV
import torch.nn.functional as F
from sklearn.model_selection import train_test_split



def load_single(path):
    sp = load_npz(path)
    X1, y1, A1, X2, y2, A2 = sp
    X1, X3, y1, y3, A1, A3 = train_test_split(X1, y1, A1, stratify=y1, test_size=0.3,
                                              random_state=0)
    print('train, test', X1.shape, X2.shape)
    sp = X1, y1, A1, X2, y2, A2, X3, y3, A3
    tensor_dataset = torch.utils.data.TensorDataset(X1, y1, A1)
    train_load = torch.utils.data.DataLoader(tensor_dataset, batch_size=256, shuffle=True)
    return {'loader': train_load, 'test': (X2, y2, A2), 'train': (X1, y1, A1),
            'verify': (X3, y3, A3)}


def load_multiply(path):
    A, B, C = load_multi_npz(path)
    X1, y1, A1 = A
    X2, y2, A2 = B
    X3, y3, A3 = C
    tensor_dataset = torch.utils.data.TensorDataset(X1, y1, A1)
    train_load = torch.utils.data.DataLoader(tensor_dataset, batch_size=256, shuffle=True)
    print('train, test', X1.shape, X2.shape)
    print('sensitive ratio', A1.mean(0))
    return {'loader': train_load, 'test': (X2, y2, A2), 'train': (X1, y1, A1),
            'verify': (X3, y3, A3)}


class EpochLoss:
    def __init__(self, flag=True, size=-1):
        self.size = size
        self.sum = []
        self.i = 1
        self.record = []
        self.met = []
        for i in range(size):
            self.sum.append(0)
            self.record.append([])
        self.f = flag

    def amlt(self, *loss):
        if self.size == -1:
            self.__init__(size=len(loss))
        for i in range(self.size):
            if torch.isnan(loss[i]):
                return False
            self.sum[i] += loss[i].item()

        return True

    def test(self, y_, y, s):
        ms = mt.Statics(y_, y, s)
        dp = abs(ms.result[0].dp[0] - ms.result[0].dp[1])
        print(f'acc = {ms.acc}, dp = {dp}')

    def message(self):
        message = []
        for i in range(self.size):
            message.append('{:0<8.5f}'.format(self.sum[i]))
            self.record[i].append(self.sum[i])
        print('\repoch {: >3d}: {}'.format(self.i, ' '.join(message)), end='')

        self.i += 1
        self.sum = []

        for i in range(self.size):
            self.sum.append(0)

    def line(self):
        print()
        if self.f:
            for i in range(self.size):
                plt.plot(self.record[i])
                plt.show()


class Result:

    def __init__(self, y_hat, y, sensitives, threshold=0.5):
        self.fair_process = None
        self.y_raw = y_hat
        self.y = y
        self.sens = {}
        self.sensitives = [sensitives[:, i] for i in range(sensitives.shape[1])]
        y_hat = np.where(y_hat > threshold, 1, 0)
        self.y_hat = y_hat
        self.caps = {}
        tp_ids = np.where(np.logical_and(y > 0, y_hat > 0))[0]
        fp_ids = np.where(np.logical_and(y < 1, y_hat > 0))[0]
        tn_ids = np.where(np.logical_and(y < 1, y_hat < 1))[0]
        fn_ids = np.where(np.logical_and(y > 0, y_hat < 1))[0]
        tp = tp_ids.shape[0]
        fp = fp_ids.shape[0]
        tn = tn_ids.shape[0]
        fn = fn_ids.shape[0]
        self.caps['all'] = [tp, fp, tn, fn]
        tpr, tnr = tp / (tp + fn), tn / (tn + fp)
        self.order = "Tpr, Fpr, Tnr, Fnr"
        ls = [tpr, 1 - tnr, tnr, 1 - tpr]
        self.raw_ = np.array(ls)
        self.raw = np.around(np.array(ls), decimals=4)
        ls = []
        r_ls = []
        acc = [np.mean(y == y_hat)]
        for i in range(sensitives.shape[1]):
            s = sensitives[:, i]
            tps = np.count_nonzero(s[tp_ids])
            fps = np.count_nonzero(s[fp_ids])
            tns = np.count_nonzero(s[tn_ids])
            fns = np.count_nonzero(s[fn_ids])
            tpr, tnr = tps / (tps + fns), tns / (tns + fps)
            ls.append([tpr, 1 - tnr, tnr, 1 - tpr])

            tpns = np.count_nonzero(s[tp_ids] == 0)
            fpns = np.count_nonzero(s[fp_ids] == 0)
            tnns = np.count_nonzero(s[tn_ids] == 0)
            fnns = np.count_nonzero(s[fn_ids] == 0)
            self.caps[i] = [[tps, fps, tns, fns], [tpns, fpns, tnns, fnns]]
            tpr, tnr = tpns / (tpns + fnns), tnns / (tnns + fpns)
            r_ls.append([tpr, 1 - tnr, tnr, 1 - tpr])

            sss, ccc = np.where(s == 0), np.where(s == 1)
            y1, y2 = y_hat[ccc], y[ccc]
            y3, y4 = y_hat[sss], y[sss]
            acc.append([np.mean(y3 == y4), np.mean(y1 == y2)])
        self.acc_ = acc
        self.acc = {'all': np.around(acc[0], decimals=4)}
        ii = 0
        for i, j in acc[1:]:
            self.acc[ii] = np.around(i, decimals=4), np.around(j, decimals=4)
            ii += 1
        ls = np.array(ls)
        r_ls = np.array(r_ls)
        self.sens_ = {i: (ls[i], r_ls[i]) for i in range(ls.shape[0])}
        ls = np.around(ls, decimals=4)
        r_ls = np.around(r_ls, decimals=4)
        self.sens = {i: (ls[i], r_ls[i]) for i in range(ls.shape[0])}

        self.group = {k: self.sens[k][0] for k in self.sens}
        self.group['all'] = self.raw

    def predict_distribution(self, sen=0):
        if self.fair_process:
            print(f'Note: S{self.fair_process} has benn fair process')
        self.fair_process = sen
        prediction_distribution(self.y_raw, self.y, self.sensitives[sen])

    def status(self):
        if self.fair_process:
            print(f'Note: S{self.fair_process} has benn fair process')
        p = pt.PrettyTable(['Group', *self.order.split(', '), 'Acc'])
        p.add_row(['all', *self.raw, self.acc['all']])
        for k in self.sens:
            mess = []
            for i, j in zip(*self.sens[k]):
                mess.append('{:0<6}  {:0<6}'.format(i, j))
            i, j = self.acc[k]
            mess.append('{:0<6}  {:0<6}'.format(i, j))

            p.add_row([f's{k}=1 s{k}=0', *mess])
        p.align = 'l'
        print(p)

    def fairness(self, p=0, s=0):
        """
        根据不同组之间TPR，FPR，TNR，FNR计算 公平 指标
        :param p:
            0: s=i 和 s!=i 之间的比较
            1: s=i 和 s=j 之间的插值最大比较
            2: s=i 和 all 之间的比较
        :param s:
        :return:
        """
        if p != 1:
            self.fair_process = s

        def fun(x, y):
            dfc = x / y if x / y < 1 else y / x
            dfc = np.around(dfc, decimals=10)
            return np.around(abs(x - y), decimals=10), dfc

        def f(x, y):
            return np.around(x, decimals=10), np.around(y, decimals=10)

        if p == 0:
            ls = ['s=0', 's=1']
            data_1, data_2 = self.caps[s]
            mess = [(j, i) for i, j in zip(*self.sens_[s])]
        elif p == 2:
            ls = ['s=all', 's=1']
            data_1 = self.caps[s][0]
            data_2 = self.caps['all']
            mess = [(j, i) for i, j in zip(self.sens_[s][0], self.raw_)]
        elif p == 1:
            u, v = s
            ls = [f's={v}', f's={u}']
            data_1 = self.caps[u][0]
            data_2 = self.caps[v][0]
            mess = [(j, i) for i, j in zip(self.sens_[u][0], self.sens_[v][0])]
        else:
            print('p取{0， 1， 2}')
            raise -1
        if self.fair_process:
            print(f'Note: S{self.fair_process} has benn fair process')
        dp_1 = (data_1[0] + data_1[1]) / sum(data_1)
        dp_2 = (data_2[0] + data_2[1]) / sum(data_2)

        res = [['Fairness Metrics', *ls, 'Diff', 'Ratio']]
        pr = pt.PrettyTable(res[0])
        res.append(['demographic parity', *f(dp_1, dp_2), *fun(dp_1, dp_2)])
        res.append(['equal opportunity', *f(*mess[0]), *fun(mess[0][1], mess[0][0])])
        x1, y1 = fun(*mess[0])
        x2, y2 = fun(*mess[2])

        if x1 > x2:
            res.append(['equal odds (max)', *f(*mess[0]), x1, y1])
        else:
            res.append(['equal odds (max)', *f(*mess[2]), x2, y2])
        for line in res[1:]:
            pr.add_row(line)
        pr.align = 'l'
        print(pr)
        res = np.array(res)
        res = res[:, 1:]
        res = res[1:, :]
        df = pd.DataFrame(res)
        df.columns = [*ls, 'Diff', 'Ratio']
        df.index = ['demographic parity', 'equal opportunity', 'equal odds (max)']

        return df

    def info(self, s=0, p=0):
        self.predict_distribution(s)
        self.fairness(s=s, p=p)
        self.status()

        self.SD()

    def violation(self):
        if self.fair_process:
            print(f'Note: S{self.fair_process} has benn fair process')
        print('*在s_i和全体之间进行violation计算')
        y_h = self.y_raw
        y_t = self.y
        ss = self.sensitives
        dps, eos, edds = [], [], []
        for s in ss:
            s_id = np.where(s > 0)
            dp = np.abs(np.mean(y_h) - np.mean(y_h[s_id]))
            positive = np.where(y_t > 0)
            pos_s_id = np.where(np.logical_and(y_t > 0, s > 0))
            eo = np.abs(np.mean(y_h[positive]) - np.mean(y_h[pos_s_id]))
            neg = np.where(y_t < 1)
            neg_s_id = np.where(np.logical_and(y_t < 1, s > 0))
            n_ = np.abs(np.mean(y_h[neg]) - np.mean(y_h[neg_s_id]))
            edd = max(eo, n_)
            dps.append(dp)
            eos.append(eo)
            edds.append(edd)
        res = np.array([dps, eos, edds]).transpose()
        res = np.insert(res, 0, np.max(res, axis=0), axis=0)
        p = pt.PrettyTable('group, dp, equal opportunity, equal odds'.split(', '))
        p.add_row(['Max', *res[0]])
        for i, line in enumerate(res[1:]):
            p.add_row([f's={i}', *line])
        print(p)

        return res

    def SD(self):
        y_h = self.y_raw
        y_t = self.y
        ss = self.sensitives
        line = []
        for s in ss:
            s_1 = np.where(s == 1)
            s_0 = np.where(s == 0)
            sdp_0 = np.mean(y_h[s_0])
            sdp_1 = np.mean(y_h[s_1])
            sdp = np.abs(np.mean(y_h[s_0]) - np.mean(y_h[s_1]))

            pos_s_1 = np.where(np.logical_and(y_t == 1, s == 1))
            pos_s_0 = np.where(np.logical_and(y_t == 1, s == 0))
            tp_d = np.abs(np.mean(y_h[pos_s_0]) - np.mean(y_h[pos_s_1]))

            neg_s_1 = np.where(np.logical_and(y_t == 0, s == 1))
            neg_s_0 = np.where(np.logical_and(y_t == 0, s == 0))
            tn_d = np.abs(np.mean(y_h[neg_s_1]) - np.mean(y_h[neg_s_0]))

            line.append([sdp, tp_d, tn_d])

        line = np.array(line)

        p = pt.PrettyTable('group, sdp, equal opportunity, equal odds'.split(', '))
        line = np.insert(line, 0, np.max(line, axis=0), axis=0)
        p.add_row(['Max', *line[0]])
        for i, item in enumerate(line[1:]):
            p.add_row([f's{i}', *item])
        print(p)

        return line


class PandasDataSet(torch.utils.data.TensorDataset):

    def __init__(self, *dataframes):
        tensors = []
        for df in dataframes:
            if isinstance(df, pd.Series):
                df = df.to_frame('dummy')
            tensors.append(torch.from_numpy(df.values).float())
        super(PandasDataSet, self).__init__(*tensors)


class VanillaModel(torch.nn.Module):
    def __init__(self, input_size, layer=(100, 50), dropout=None):
        super(VanillaModel, self).__init__()
        if dropout:
            p_dropout = dropout
            self.network = torch.nn.Sequential(
                torch.nn.Linear(input_size, layer[0]),
                torch.nn.ReLU(),
                torch.nn.Dropout(p_dropout),
                torch.nn.Linear(layer[0], layer[1]),
                torch.nn.ReLU(),
                torch.nn.Dropout(p_dropout),
                torch.nn.Linear(layer[1], 2)
            )
        else:
            self.network = torch.nn.Sequential(
                torch.nn.Linear(input_size, layer[0]),
                torch.nn.ReLU(),
                torch.nn.Linear(layer[0], layer[1]),
                torch.nn.ReLU(),
                torch.nn.Linear(layer[1], 2)
            )

    def forward(self, x):
        return torch.softmax(self.network(x), dim=1)


def seed_everything(seed):
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    random.seed(seed)
    np.random.seed(seed)


def mmd(source, target, kernel_mul=2.0, kernel_num=5, fix_sigma=None):
    """
    Max-min-distance
    :param source: torch.Tensor 2维
    :param target: torch.Tensor 2维
    :param kernel_mul:
    :param kernel_num:
    :param fix_sigma:
    :return:
    """
    return mt.mmd(source, target, kernel_mul, kernel_num, fix_sigma)


"""
load dataset
"""


def dataset_message(data):
    X, y, S = data
    print(f'sample num: {X.shape[0]}, feature: {X.shape[1]}, {S.shape[1]}')
    print(f'sensitive: {S.columns}')
    return X, y, S


def load_dataset(name, path=None):
    # if path is None:
    #     path = f'dataset/{name}/{name}.csv'
    data = None
    if name == 'adult':
        data = load_adult(path)
    elif name == 'compas':
        data = load_compas(path)
    elif name == 'bank':
        data = load_bank(path)
    else:
        raise NotImplemented(f'no this dataset: {name}')
    return data


def load_bank(path=None):
    pjk = dlt.load_bank_(path)
    return dataset_message(pjk)


def load_adult(path=None):
    pjk = dlt.load_adult(path)
    return dataset_message(pjk)


def load_compas(path=None):
    pjk = dlt.load_compas(path)
    return dataset_message(pjk)


def load_data(path):
    column_names = ['age', 'workclass', 'fnlwgt', 'education', 'education_num',
                    'martial_status', 'occupation', 'relationship', 'race', 'sex',
                    'capital_gain', 'capital_loss', 'hours_per_week', 'country', 'target']
    input_data = (pd.read_csv(path, names=column_names,
                              na_values="?", sep=r'\s*,\s*', engine='python')
        .loc[lambda df: df['race'].isin(['White', 'Black'])])

    print(f'原始表数据：{input_data.shape}')
    sensitive_attribs = ['race', 'sex']
    Z = (input_data.loc[:, sensitive_attribs]
         .assign(race=lambda df: (df['race'] == 'White').astype(int),
                 sex=lambda df: (df['sex'] == 'Male').astype(int)))

    y = (input_data['target'] == '>50K').astype(int)

    X = (input_data
         .drop(columns=['target', 'fnlwgt'])
         .fillna('Unknown')
         .pipe(pd.get_dummies, drop_first=True))
    scaler = StandardScaler().fit(X)
    scale_df = lambda df, scaler: pd.DataFrame(scaler.transform(df),
                                               columns=df.columns, index=df.index)
    X = X.pipe(scale_df, scaler)

    return X, y, Z


def draw_ROC(t_label, predict):
    fpr, tpr, thresholds = metrics.roc_curve(t_label, predict)
    auc = metrics.auc(fpr, tpr)
    plt.plot(fpr, tpr, label=f'AUC:{auc:.3f}')
    plt.legend(loc='lower right')
    plt.show()

    return find_threshold(fpr, tpr, thresholds), auc


def softmax(outputs):
    return torch.softmax(outputs, dim=-1)


def find_threshold(fpr, tpr, thresholds):
    ls = [(a, b) for a, b in zip(fpr, tpr)]
    ls.sort(key=lambda x: x[1] - x[0])
    fpr = list(fpr)
    return thresholds[fpr.index(ls[-1][0])]


def demographic_parity(predict_group):
    """
    :param predict_group:
    :return:
    """
    a_0, a_1 = predict_group
    k1, k2 = a_0.mean() / a_1.mean(), a_1.mean() / a_0.mean()
    return min(k1, k2)


def save_log(path, logger):
    with open('log/{}.txt'.format(path), 'w', encoding='utf-8') as f:
        f.write("\n".join(logger))


def analyze(predict, label):
    """
    :param predict: 预测的类别
    :param label: 真实的类别
    :return:
    """
    mat = metrics.confusion_matrix(label, predict)
    f1 = metrics.f1_score(label, predict)
    fpr = mat[0][1] / sum(mat[0])
    fnr = mat[1][0] / sum(mat[1])
    acc = metrics.accuracy_score(label, predict)
    auc = metrics.roc_auc_score(label, predict)
    return {'acc': acc, 'f1': f1, 'fpr': fpr, 'fnr': fnr, 'auc': auc}


def sensitive_in(predict, label, sensitive):
    """
    用于数据分组
    :param predict:  预测值
    :param label:  标签值
    :param sensitive: 敏感属性（1维）
    :return: sensitive_dic:字典，key是敏感属性类别，值是[[], []]，其中第一个列表是标签值，第二个是预测值
    """
    sensitive_dic = {}
    for p, l, a, in zip(predict, label, sensitive):
        sensitive_dic[a] = sensitive_dic.get(a, [[], []])
        sensitive_dic[a][0].append(l)
        sensitive_dic[a][1].append(p)
    return sensitive_dic


import prettytable as pt
from sklearn.decomposition import PCA
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def sample_2D_distribution(raw, sensitive):
    zz_2 = PCA(2).fit_transform(raw)
    g1, g2 = [], []
    for z, a in zip(zz_2, sensitive):
        if a == 1:
            g1.append(z.tolist())
        else:
            g2.append(z.tolist())
    g1, g2 = np.array(g1), np.array(g2)
    plt.figure(dpi=80)
    plt.scatter(x=g2[:, 0], y=g2[:, 1], s=80, alpha=0.2, label='female')
    plt.scatter(x=g1[:, 0], y=g1[:, 1], s=80, c='pink', alpha=0.2, label='male')
    plt.legend()
    plt.show()


def save(path, dt, note, model):
    m1, m2 = dt
    ls1 = []
    ls2 = []
    for k in m1:
        data = [k] + m1[k]
        ls1.append(data)
    for k in m2:
        data = [k] + m2[k]
        ls2.append(data)
    orange_fill = PatternFill(fill_type='solid', fgColor="FFC125")
    wb = load_workbook(path)
    sheet = wb['result']
    start_row = sheet.max_row + 2

    sheet.cell(start_row, 1).value = note
    sheet.cell(start_row, 2).value = model
    start_row += 1
    lines = ['item', 'S0', 'S1', 'total', 'ratio']
    for i in range(len(lines)):
        sheet.cell(start_row, i + 1).value = lines[i]
    start_row += 1
    for r, line in enumerate(ls1):
        for i in range(len(line)):
            if r == 1 and i == 4:
                sheet.cell(r + start_row, i + 1).fill = orange_fill
            sheet.cell(r + start_row, i + 1).value = line[i]
    start_row += len(ls1)
    start_row += 1
    lines = ['fairness metric', 'value']
    for i in range(len(lines)):
        sheet.cell(start_row, i + 1).value = lines[i]
    start_row += 1
    for r, line in enumerate(ls2):
        last = sheet.max_row + 1
        for i in range(len(line)):
            if i == 1:
                sheet.cell(r + start_row, i + 1).fill = orange_fill
            sheet.cell(r + start_row, i + 1).value = line[i]
    wb.save(path)


def performance(prob, label, sensitive):
    res, f = binary([1 if i > 0.5 else 0 for i in prob], label, sensitive)
    tb = pt.PrettyTable()
    tb.field_names = ["Item", "Z0", "Z1", "total", "Z0/Z1"]
    for v in res:
        tb.add_row([v, *res[v]])
    tb.align = 'l'
    print(tb)

    ls = []
    for k in f:
        ls.append([k, *f[k]])
    tb2 = pt.PrettyTable(['fairness Metric', 'value'])
    # tb.header = False
    for line in ls:
        tb2.add_row(line)
    tb2.align = 'l'
    print(tb2)
    return res, f


"""
**************
* plot tools *
**************
"""


def discrepancy(X, cat):
    """
    use sample
    discrepancy(Z2.numpy(), A2[:, 1].numpy())
    where Z2 is the representation layer
    A2 is the sensitive features
    :param X: data
    :param cat:  sensitive list like [1, 0, 1, ...]
    :return:
    """
    ids = np.where(cat > 0)
    d_1 = X[ids]
    ids = np.where(cat < 1)
    d_0 = X[ids]
    plt.figure(dpi=100, figsize=(20, 15))
    for i in range(X.shape[1]):
        plt.subplot(X.shape[1] // 3 + 1, 3, i + 1)
        plt.title(f'z{i}')
        sns.kdeplot(d_0[:, i], shade=True)
        sns.kdeplot(d_1[:, i], shade=True)
    plt.show()


def data_distribution(raw, cat):
    """
    sample:
    data_distribution(Z2.numpy()[:, 2:], A2[:, 1].numpy())
    data_distribution(Z2.numpy()[:, :-2], A2[:, 1].numpy())

    show the diff between s=0 and s=1 in raw data
    :param raw:
    :param cat:  sensitive list like [1, 0, 1, ...]
    :return:
    """
    data = PCA(2).fit_transform(raw)
    ids = np.where(cat > 0)
    d_1 = data[ids]
    ids = np.where(cat < 1)
    d_0 = data[ids]
    plt.figure(dpi=80)
    plt.scatter(x=d_1[:, 0], y=d_1[:, 1], s=80, alpha=0.2, label='s=1')
    plt.scatter(x=d_0[:, 0], y=d_0[:, 1], s=80, c='pink', alpha=0.2, label='s=0')
    plt.legend()
    plt.show()


def prediction_distribution(prob, label, sensitive, title='{}'):
    prob_0 = []
    prob_1 = []
    for p, l, s in zip(prob, label, sensitive):
        if s == 0:
            prob_0.append(p)
        else:
            prob_1.append(p)
    sns.set()
    plt.figure()
    sns.kdeplot(prob_0, shade=True, label='Z0')
    sns.kdeplot(prob_1, shade=True, label='Z1')
    plt.title(title.format('predict distributions'), fontsize='15')
    plt.xlim(0, 1)
    plt.ylim(0, 7)
    plt.yticks([])
    plt.ylabel('Prediction distribution')
    plt.show()
    return prob_0, prob_1


def binary(predict, label, sensitive):
    check = set(sensitive)
    if check != {0, 1}:
        print('敏感属性值只能取0或1，且必须包括0， 1')
        return None
    predict_0, label_0 = [], []
    predict_1, label_1 = [], []
    p_0, p_1, p = 0, 0, 0
    n_0, n_1, n = 0, 0, 0
    for p, l, s in zip(predict, label, sensitive):
        if s == 0:
            predict_0.append(p)
            label_0.append(l)
            if l == 1:
                p_0 += 1
                p += 1
            else:
                n_0 += 1
                n += 1
        else:
            predict_1.append(p)
            label_1.append(l)
            if l == 1:
                p_1 += 1
                p += 1
            else:
                n_1 += 1
                n += 1
    dp_0 = 0
    for h in predict_0:
        if h == 1:
            dp_0 += 1
    dp_0 = dp_0 / len(predict_0)
    dp_1 = 0
    for h in predict_1:
        if h == 1:
            dp_1 += 1
    dp_1 = dp_1 / len(predict_1)
    try:
        dp = dp_0 / dp_1
        dp = min(dp, 1 / dp)
    except ZeroDivisionError:
        dp = '-'
    acc_0 = metrics.accuracy_score(label_0, predict_0)
    acc_1 = metrics.accuracy_score(label_1, predict_1)
    acc = metrics.accuracy_score(label, predict)

    auc_0 = metrics.roc_auc_score(label_0, predict_0)
    auc_1 = metrics.roc_auc_score(label_1, predict_1)
    auc = metrics.roc_auc_score(label, predict)
    mat_0 = metrics.confusion_matrix(label_0, predict_0, labels=[0, 1])
    mat_1 = metrics.confusion_matrix(label_1, predict_1, labels=[0, 1])
    mat = metrics.confusion_matrix(label, predict, labels=[0, 1])

    fpr_0 = mat_0[0][1] / sum(mat_0[0])
    fnr_0 = mat_0[1][0] / sum(mat_0[1])
    fpr_1 = mat_1[0][1] / sum(mat_1[0])
    fnr_1 = mat_1[1][0] / sum(mat_1[1])
    fpr = mat[0][1] / sum(mat[0])
    fnr = mat[1][0] / sum(mat[1])

    result = {'acc': [acc_0, acc_1, acc, ratio(acc_0, acc_1)], 'dp': [dp_0, dp_1, abs(dp_1 - dp_0), dp],
              'fpr': [fpr_0, fpr_1, fpr, ratio(fpr_0, fpr_1)], 'fnr': [fnr_0, fnr_1, fnr, ratio(fnr_0, fnr_1)],
              'tpr': [1 - fnr_0, 1 - fnr_1, 1 - fnr, ratio(1 - fnr_0, 1 - fnr_1)],
              'auc': [auc_0, auc_1, auc, ratio(auc_0, auc_1)]}
    fairness = {'equal opportunity  p(h=1|y=1,z)=p(h=1|y=1)': [abs(fnr_0 - fnr_1)],
                'demographic parity p(h=1|z)=p(h=1)': [abs(dp_0 - dp_1)],
                'equal odds max(tpr, tnr)': [max(abs(fpr_0 - fpr_1), abs(fnr_0 - fnr_1))]
                }
    return result, fairness


def ratio(x1, x2):
    k = x1 / x2
    if k > 1:
        k = 1 / k
    return k


def show(predict, label, sensitive, sname, class_gap=0.5, epoch=0):
    """
    数据可视化~
    :param predict: 预测的类别概率
    :param label: 真实的类别
    :param sensitive: 敏感属性（1维）
    :return:
    如 show([0.6, 0.1], [1, 0], [1, 1])
    """
    if type(predict) != torch.Tensor:
        predict = torch.Tensor(predict)
    predict = predict.view(-1)
    sensitive_values = [0, 1]

    predict_ = [predict[sensitive[sname].values == i] for i in sensitive_values]
    label_ = [label[sensitive[sname].values == i] for i in sensitive_values]
    pre_class = []
    for group in sensitive_values:
        pre_class.append(pd.Series([1 if i > class_gap else 0 for i in predict_[group]]))
    d = []
    fpr_fnr = []
    for group in sensitive_values:
        h = analyze(pre_class[group], label_[group])
        fpr_fnr.append([h.get('fpr'), h.get('fnr')])
        d.append(h)

    text = [f'demographic parity:{demographic_parity(pre_class) * 100:.2f}%']

    sns.set()
    plt.figure(figsize=(6, 4))
    plt.title(f'#{epoch} predict distributions for {sname}+ & {sname}-', fontsize='15')
    plt.xlim(0, 1)
    plt.ylim(0, 7)
    plt.yticks([])
    plt.ylabel('Prediction distribution')
    for k in sensitive_values:
        sns.kdeplot(predict_[k].tolist(), shade=True, label=k)
        word = [f'{kk}={d[k][kk]:.3f}' for kk in d[k]]
        text.append(f'{k:} : ' + ', '.join(word))
    plt.text(1.1, 6, '\n'.join(text), fontsize='16')


attribute_indexes = {'age': [0, 1],
                     'education_num': [1, 2],
                     'capital_gain': [2, 3],
                     'capital_loss': [3, 4],
                     'hours_per_week': [4, 5],
                     'workclass': [5, 13],
                     'education': [13, 28],
                     'martial_status': [28, 34],
                     'occupation': [34, 48],
                     'relationship': [48, 53],
                     'country': [53, 93]}


def analyze_distribution_scatter(X, y, Z, attribute, sensitive):
    l, r = attribute_indexes.get(attribute)
    result0 = []
    result1 = []
    for i in range(l, r):
        result0.append(0)
        result1.append(0)
    result0 = np.asarray(result0, dtype=float)
    result1 = np.asarray(result1, dtype=float)
    size = len(X)
    z0 = 0
    z1 = 0
    for i in range(size):
        yy = y[i]
        z = Z[i]
        if yy == 1:
            continue
        if z == 0:
            z0 += 1
            result0 += X[i, l:r].numpy()
        else:
            z1 += 1
            result1 += X[i, l:r].numpy()
    result0 = result0 / z0
    result1 = result1 / z1

    plt.title(f'{sensitive} to {attribute}')
    plt.bar(range(len(result0)), result0, label=f'{sensitive}0')
    plt.bar(range(len(result1)), -result1, label=f'{sensitive}1')
    plt.legend(loc='lower right')
    plt.savefig(f'outputs/dataset_n/{sensitive}-{attribute}.png')
    plt.show()


def p_rule(y_pred, z_values, threshold=0.5):
    y_z_1 = y_pred[z_values == 1] > threshold
    y_z_0 = y_pred[z_values == 0] > threshold
    odds = y_z_1.mean() / y_z_0.mean()
    return np.min([odds, 1 / odds]) * 100


def plot_distributions(y_true, Z_true, y_pred, Z_pred=None, epoch=None):
    fig, axes = plt.subplots(1, 2, figsize=(10, 4))

    subplot_df = (
        Z_true
            .assign(race=lambda x: x['race'].map({1: 'white', 0: 'black'}))
            .assign(sex=lambda x: x['sex'].map({1: 'male', 0: 'female'}))
            .assign(y_pred=y_pred)
    )
    _subplot(subplot_df, 'race', ax=axes[0])
    _subplot(subplot_df, 'sex', ax=axes[1])
    _performance_text(fig, y_true, Z_true, y_pred, Z_pred, epoch)
    fig.tight_layout()
    return fig


def corr(x, y):
    """
    相关系数 越低，越不相关
    """
    xm, ym = torch.mean(x), torch.mean(x)
    xvar = torch.sum((x - xm) ** 2) / x.shape[0]
    yvar = torch.sum((y - ym) ** 2) / x.shape[0]
    return torch.abs(torch.sum((x - xm) * (y - ym)) / (xvar * yvar) ** 0.5)


def _subplot(subplot_df, col, ax):
    for label, df in subplot_df.groupby(col):
        sns.kdeplot(df['y_pred'], ax=ax, label=label, shade=True)
    ax.set_title(f'Sensitive attribute: {col}')
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 7)
    ax.set_yticks([])
    ax.set_ylabel('Prediction distribution')
    ax.set_xlabel(r'$P({{income>50K}}|z_{{{}}})$'.format(col))


def _performance_text(fig, y_test, Z_test, y_pred, Z_pred=None, epoch=None):
    if epoch is not None:
        fig.text(1.0, 0.9, f"Training epoch #{epoch}", fontsize='16')

    clf_roc_auc = metrics.roc_auc_score(y_test, y_pred)
    clf_accuracy = metrics.accuracy_score(y_test, y_pred > 0.5) * 100
    p_rules = {'race': p_rule(y_pred, Z_test['race']),
               'sex': p_rule(y_pred, Z_test['sex']), }
    fig.text(1.0, 0.65, '\n'.join(["Classifier performance:",
                                   f"- ROC AUC: {clf_roc_auc:.2f}",
                                   f"- Accuracy: {clf_accuracy:.1f}"]),
             fontsize='16')
    fig.text(1.0, 0.4, '\n'.join(["Satisfied p%-rules:"] +
                                 [f"- {attr}: {p_rules[attr]:.0f}%-rule"
                                  for attr in p_rules.keys()]),
             fontsize='16')
    if Z_pred is not None:
        adv_roc_auc = metrics.roc_auc_score(Z_test, Z_pred)
        fig.text(1.0, 0.20, '\n'.join(["Adversary performance:",
                                       f"- ROC AUC: {adv_roc_auc:.2f}"]),
                 fontsize='16')


def load_npz(npz_file):
    data = np.load(npz_file)
    res = {}
    for k in data:
        res[k] = data[k]
    aps = [('attr_train', 's_train'), ('attr_test', 's_test')]
    for i, j in aps:
        if i in res and j in res:
            continue
        if i in res:
            res[j] = res[i]
        elif j in data:
            res[i] = res[j]
        else:
            continue
    X1 = res['x_train']
    y1 = res['y_train']
    A1 = res['attr_train']
    X2 = res['x_test']
    y2 = res['y_test']
    A2 = res['attr_test']
    print(f'x shape:{X1.shape} y shape:{y1.shape} s shape:{A1.shape}')
    print('x columns ', res['input_columns'])
    print('s columns ', res['s_columns'])
    return list(map(lambda x: torch.from_numpy(x).float(), [X1, y1, A1, X2, y2, A2]))


def load_multi_npz(file_path):
    data = np.load(file_path)
    res = {}
    for k in data:
        res[k] = data[k]
    pre = 'x,y,s'.split(',')
    post = 'train test valid'.split(' ')
    out = []
    for item in post:
        for name in pre:
            k = f'{name}_{item}'
            out.append(torch.from_numpy(res[k]).float())

    print('x columns ', res['input_columns'])
    print('s columns ', res['s_columns'])
    train, test, valid = out[:3], out[3:6], out[6:]
    return train, test, valid


from methods import FNNC, Adv, CFair, Net, FRC, AdvS, Corr


def parameters_build(args, d):
    class_name = args.method
    p = args.parameters
    names = ['fnnc', 'adv', 'cfair', 'net', 'frc', 'ads', 'corr']
    if class_name not in names:
        raise NotImplemented(f'method should in {names}')
    if class_name == 'fnnc':
        ans = {'fnnc_gamma': p[0],
               'fnnc_loss_method': args.criterion}
    elif class_name == 'adv':
        ans = {'mu': p[0]}
    elif class_name == 'corr':
        ans = {'corr_gamma': p[0]}
    elif class_name == 'cfair':
        ans = {'mu': p[0]}
    elif class_name == 'frc':
        ans = {'parameters': p, 'z_dim': args.z_dim}

        if args.dataset == 'compas':
            ans['act'] = 'sigmoid'

    elif class_name == 'ads':
        ans = {'parameters': p}
    elif class_name == 'net':
        ans = {}
    else:
        raise NotImplemented(f'method should in {names}')
    ans['device'] = d
    return ans


def method_build(class_name):
    models = [FNNC.FNNC, Adv.FairNet, CFair.CFairNet, Net.Net, FRC.FRC, AdvS.Adversarial,
              Corr.Corr]
    names = ['fnnc', 'adv', 'cfair', 'net', 'frc', 'ads', 'corr']
    if class_name not in names:
        raise NotImplemented(f'method should in {names}')
    return {i: j for i, j in zip(names, models)}.get(class_name)
